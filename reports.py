import pandas as pd
import openpyxl

def save_to_csv(articles, filename="articles.csv"):
    df = pd.DataFrame(articles) #converts the articles to a DataFrame
    df.to_csv(filename, index=False) #saves the DataFrame to a CSV file
    print(f"Results saved to {filename}")

def save_to_excel(articles, filename="articles.xlsx"):
    
    workbook = openpyxl.Workbook() #creates a new Excel workbook
    worksheet = workbook.active #gets the active worksheet
    worksheet.title = "Articles" #sets the title of the worksheet

    headers = ["Authors", "Year of Publication", "APA-style Citation", "Abstract", "DOI", "Number of Citations"] #creates a list of headers
    worksheet.append(headers) #appends the headers to the worksheet

    for row, article in enumerate(articles, start=2): #iterates over the articles
        authors = article.get("authors", "Unknown") #extracts the authors from the article
        if isinstance(authors, list): #if the authors are a list of strings,
            authors = ", ".join(authors) #converts the authors to a string
        
        #writes the article information to the worksheet
        worksheet.cell(row=row, column=1, value=authors)
        worksheet.cell(row=row, column=2, value=article.get("pub_year", "Unknown"))
        worksheet.cell(row=row, column=3, value=article.get("apa-style_citation", ""))
        worksheet.cell(row=row, column=4, value=article.get("abstract", "No abstract available"))
        worksheet.cell(row=row, column=5, value=article.get("DOI", "No DOI"))
        worksheet.cell(row=row, column=6, value=article.get("num_citations", 0))

    col_letters = ["A", "B", "C", "D", "E", "F"] #creates a list of column letters
    i = 0 #initializes the column index
    for col in worksheet.columns: #iterates over the columns
        max_length = 0 #initializes the maximum length of the column
        column = col_letters[i] #gets the column letter
        for cell in col: #iterates over the cells in the column
            if max_length < len(str(cell.value)): #if the length of the cell value is greater than the maximum length,
                max_length = len(str(cell.value)) #updates the maximum length
        worksheet.column_dimensions[column].width = max_length  #sets the column width based on the maximum length
        i+=1 #increments the column index

    workbook.save(filename) #saves the workbook to a file

    